"""Roadmap Step 1 -- Excel Mapping Loader.

The ONE place in the backend that touches the therapy-mapping `.xlsx` files. Its
sole job is to locate, open, validate, and load the two mapping workbooks
(Neurodivergent_map.xlsx / Neurotypical_map.xlsx) under `data/roadmap/` into
immutable in-memory datasets, and return them.

It does NOT map domains to therapies, search, filter, forward-fill merged cells,
rank, or generate a roadmap -- those build FROM this loader's output in later
modules. Every value is carried through verbatim (spec S3.3).

Not to be confused with `roadmap_loader.py`, which loads the frontend JSON
submission -- a different concern entirely.

Public surface
--------------
- ``load_mappings(base_dir=None, *, strict=True)`` -- fresh load. ``base_dir=None``
  uses ``settings.roadmap_mapping_dir``; an explicit ``base_dir`` (e.g. a test
  ``tmp_path``) is loaded in isolation and never touches the module cache.
- ``get_mappings(force_reload=False)`` -- cache-aware accessor over the configured
  directory. The cache fills lazily on first call -- there is deliberately NO
  eager import-time load, because this loader raises on a bad file and a
  crash-at-import would take down app startup.
- ``reload_mappings()`` -- force a cache refresh.

Any failure raises ``MappingLoadError`` (and only that); the loader never returns
a partial bundle.
"""

import logging
import os
import re
from pathlib import Path
from typing import Optional
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from app.config import settings
from app.roadmap.mapping_models import MappingBundle, MappingDataset, ModalityRow

logger = logging.getLogger("app.roadmap.mapping_loader")


class MappingLoadError(Exception):
    """The one and only error the loader raises out of ``load_mappings``.

    Carries a machine-readable ``code`` (see the table in spec S9), a human
    ``message``, the ``source`` file identity (``"ND"``/``"NT"``) when one
    applies, and the offending ``field`` (column / sheet / path) when one
    applies. Mirrors ``RoadmapValidationError`` in ``roadmap_loader.py``.
    """

    def __init__(
        self,
        code: str,
        message: str,
        source: Optional[str] = None,
        field: Optional[str] = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.source = source
        self.field = field


# --------------------------------------------------------------------------- #
# Declared-as-data knobs (Open/Closed -- extending these needs no logic edit).
# --------------------------------------------------------------------------- #

REQUIRED_COLUMNS = ("Domain", "Track", "Suggested modalities (in order)", "Relevance")

# The (identity, filename, expected worksheet) of each workbook. The observed
# sheet names differ per file, so the expected name is data, not a hardcode.
_FILES = (
    ("ND", "Neurodivergent_map.xlsx", "Neurodivergent Path"),
    ("NT", "Neurotypical_map.xlsx", "Neurotypical Path"),
)

# How many rows from the top we look through for the header. The real files put
# it at sheet row 4 (a title banner + two blank rows sit above it), so 20 is
# generous headroom.
HEADER_SCAN_LIMIT = 20

# A candidate header row must match at least this many required columns. Keeps a
# stray data row that happens to contain one column word from being mistaken for
# the header, while still letting a header that is MISSING a column be found (so
# we can raise `missing_column` instead of a vague `header_not_found`).
HEADER_MIN_MATCHES = 2

# After the last content row, this many consecutive fully-blank rows end the
# scan. This is what stops the read-only iterator from walking the ND sheet's
# 1,048,576-row reported dimension.
MAX_TRAILING_BLANK_ROWS = 50

# Cached-value error sentinels openpyxl surfaces for a cell whose formula errored
# (with data_only=True). A modality that is one of these is unrecoverable.
_EXCEL_ERROR_LITERALS = frozenset(
    {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!", "#SPILL!", "#GETTING_DATA"}
)

_REQUIRED_NORM = None  # populated lazily below


# --------------------------------------------------------------------------- #
# Local, throwaway normalization -- used ONLY to match header cells. It is never
# written into any returned value (spec S3.3 / S7.2).
# --------------------------------------------------------------------------- #


def _norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value).strip()).casefold()


def _required_norm() -> dict[str, str]:
    """normalized-header -> canonical required column name."""
    global _REQUIRED_NORM
    if _REQUIRED_NORM is None:
        _REQUIRED_NORM = {_norm(col): col for col in REQUIRED_COLUMNS}
    return _REQUIRED_NORM


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


# --------------------------------------------------------------------------- #
# Per-file loading.
# --------------------------------------------------------------------------- #


def _resolve_columns(header_values: tuple, source: str) -> dict[str, int]:
    """Given the discovered header row, return canonical-column -> cell index.

    Raises `duplicate_column` (two cells map to one required column) or
    `missing_column` (a required column is absent).
    """
    required_norm = _required_norm()
    resolved: dict[str, int] = {}
    seen_extra: list[str] = []
    for idx, cell in enumerate(header_values):
        if _is_blank(cell):
            continue
        canonical = required_norm.get(_norm(cell))
        if canonical is None:
            seen_extra.append(str(cell))
            continue
        if canonical in resolved:
            raise MappingLoadError(
                "duplicate_column",
                f"Column {canonical!r} appears more than once in the header.",
                source=source,
                field=canonical,
            )
        resolved[canonical] = idx

    missing = [col for col in REQUIRED_COLUMNS if col not in resolved]
    if missing:
        raise MappingLoadError(
            "missing_column",
            f"Header is missing required column(s): {', '.join(missing)}.",
            source=source,
            field=missing[0],
        )
    if seen_extra:
        logger.info("[%s] extra columns ignored: %s", source, seen_extra)
    return resolved


def _read_sheet(ws, source: str, strict: bool) -> tuple[int, list[ModalityRow], list[str]]:
    """Stream one worksheet into rows.

    Returns (header_row_number, rows, warnings). Raises on structural problems
    (`empty_worksheet`, `header_not_found`, column errors, `no_data_rows`, and --
    in strict mode -- `invalid_cell_type`).
    """
    required_norm = _required_norm()
    row_iter = ws.iter_rows(values_only=True)

    # ---- Phase 1: discover the header row (bounded by HEADER_SCAN_LIMIT). ----
    columns: Optional[dict[str, int]] = None
    header_row_num = 0
    saw_any_nonblank = False
    current_row = 0
    for values in row_iter:
        current_row += 1
        if current_row > HEADER_SCAN_LIMIT:
            break
        if values is None or all(_is_blank(c) for c in values):
            continue
        saw_any_nonblank = True
        matches = sum(
            1 for c in values if not _is_blank(c) and _norm(c) in required_norm
        )
        if matches >= HEADER_MIN_MATCHES:
            columns = _resolve_columns(values, source)  # raises on missing/duplicate
            header_row_num = current_row
            break

    if columns is None:
        if not saw_any_nonblank:
            raise MappingLoadError(
                "empty_worksheet",
                f"Worksheet {ws.title!r} has no non-empty rows.",
                source=source,
            )
        raise MappingLoadError(
            "header_not_found",
            f"No header row with the required columns was found in the first "
            f"{HEADER_SCAN_LIMIT} rows of {ws.title!r}.",
            source=source,
        )

    i_domain = columns["Domain"]
    i_track = columns["Track"]
    i_modality = columns["Suggested modalities (in order)"]
    i_relevance = columns["Relevance"]

    def _cell(values: tuple, idx: int) -> object:
        return values[idx] if idx < len(values) else None

    # ---- Phase 2: stream data rows (blank-row sentinel bounds the walk). ----
    rows: list[ModalityRow] = []
    warnings: list[str] = []
    consecutive_blank = 0
    for values in row_iter:
        current_row += 1
        if values is None:
            values = ()

        domain = _cell(values, i_domain)
        track = _cell(values, i_track)
        modality = _cell(values, i_modality)
        relevance = _cell(values, i_relevance)

        if all(_is_blank(v) for v in (domain, track, modality, relevance)):
            consecutive_blank += 1
            if consecutive_blank >= MAX_TRAILING_BLANK_ROWS:
                break
            continue
        consecutive_blank = 0

        # An error sentinel in the modality cell -> the row's therapy is lost.
        if isinstance(modality, str) and modality.strip() in _EXCEL_ERROR_LITERALS:
            msg = f"row {current_row}: modality cell holds Excel error {modality.strip()!r}"
            if strict:
                raise MappingLoadError("invalid_cell_type", msg.capitalize(), source=source)
            warnings.append(msg)
            logger.warning("[%s] dropped %s", source, msg)
            continue

        # A row with no modality is not a content row (e.g. a stray note); skip
        # it without letting it trip the trailing-blank sentinel.
        if _is_blank(modality):
            msg = f"row {current_row}: non-blank row without a modality -- skipped"
            warnings.append(msg)
            logger.debug("[%s] %s", source, msg)
            continue

        # Values pass through VERBATIM. A non-str, non-None value in a text column
        # is a data anomaly -> hard error (strict) or drop (non-strict); we never
        # silently coerce it to a string.
        try:
            rows.append(
                ModalityRow(
                    domain=domain,
                    track=track,
                    modality=modality,
                    relevance=relevance,
                    source_row=current_row,
                )
            )
        except Exception as exc:  # pydantic ValidationError -> typed error
            msg = f"row {current_row}: unexpected cell type ({exc.__class__.__name__})"
            if strict:
                raise MappingLoadError("invalid_cell_type", msg, source=source) from exc
            warnings.append(msg)
            logger.warning("[%s] dropped %s", source, msg)

    if not rows:
        raise MappingLoadError(
            "no_data_rows",
            f"Worksheet {ws.title!r} has a valid header but no data rows.",
            source=source,
        )

    return header_row_num, rows, warnings


def _select_sheet(wb, expected_sheet: str, source: str):
    """Pick the worksheet: expected name (case-insensitive), else the sole sheet,
    else `missing_worksheet`."""
    by_norm = {name.strip().casefold(): name for name in wb.sheetnames}
    match = by_norm.get(expected_sheet.strip().casefold())
    if match is not None:
        return wb[match]
    if len(wb.sheetnames) == 1:
        return wb[wb.sheetnames[0]]
    raise MappingLoadError(
        "missing_worksheet",
        f"Expected worksheet {expected_sheet!r} not found; sheets present: "
        f"{wb.sheetnames}.",
        source=source,
        field=expected_sheet,
    )


def _load_one(path: Path, source: str, expected_sheet: str, strict: bool) -> MappingDataset:
    # ---- File-level checks (before opening). ----
    if not path.exists():
        raise MappingLoadError(
            "file_not_found", f"Mapping file not found: {path}", source=source, field=str(path)
        )
    if not path.is_file() or path.suffix.lower() != ".xlsx":
        raise MappingLoadError(
            "not_a_file",
            f"Mapping path is not a readable .xlsx file: {path}",
            source=source,
            field=str(path),
        )
    if not os.access(path, os.R_OK):
        raise MappingLoadError(
            "permission_denied", f"Mapping file is not readable: {path}", source=source, field=str(path)
        )

    # ---- Open + read (workbook always closed, even on the error path). ----
    wb = None
    try:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except PermissionError as exc:
            raise MappingLoadError(
                "permission_denied", f"Permission denied opening {path}: {exc}", source=source
            ) from exc
        except (BadZipFile, InvalidFileException, KeyError, OSError) as exc:
            raise MappingLoadError(
                "workbook_unreadable",
                f"Workbook could not be opened (corrupt or not a real .xlsx): {path} ({exc})",
                source=source,
            ) from exc

        ws = _select_sheet(wb, expected_sheet, source)
        header_row, rows, warnings = _read_sheet(ws, source, strict)
    finally:
        if wb is not None:
            wb.close()

    logger.info(
        "[%s] workbook loaded: sheet=%r header_row=%d rows=%d warnings=%d",
        source,
        ws.title,
        header_row,
        len(rows),
        len(warnings),
    )
    return MappingDataset(
        source=source,
        path=str(path),
        sheet_name=ws.title,
        header_row=header_row,
        rows=tuple(rows),
        warnings=tuple(warnings),
    )


def _load_bundle(base_dir: Path, strict: bool) -> MappingBundle:
    logger.info("mapping load started: base_dir=%s strict=%s", base_dir, strict)
    datasets: dict[str, MappingDataset] = {}
    for source, filename, expected_sheet in _FILES:
        path = base_dir / filename
        logger.debug("resolved %s -> %s", source, path)
        try:
            datasets[source] = _load_one(path, source, expected_sheet, strict)
        except MappingLoadError as exc:
            logger.error(
                "mapping load failed: code=%s source=%s field=%s: %s",
                exc.code,
                exc.source,
                exc.field,
                exc.message,
            )
            raise

    bundle = MappingBundle(neurodivergent=datasets["ND"], neurotypical=datasets["NT"])
    logger.info(
        "mapping load complete: ND=%d rows, NT=%d rows",
        bundle.neurodivergent.row_count,
        bundle.neurotypical.row_count,
    )
    return bundle


# --------------------------------------------------------------------------- #
# Public API.
# --------------------------------------------------------------------------- #

_CACHE: Optional[MappingBundle] = None


def load_mappings(base_dir: object = None, *, strict: bool = True) -> MappingBundle:
    """Load both mapping workbooks fresh and return a `MappingBundle`.

    `base_dir=None` uses `settings.roadmap_mapping_dir`. An explicit `base_dir`
    (str or Path) is loaded in isolation and never reads or writes the module
    cache -- this is what lets tests point at a `tmp_path`. Raises
    `MappingLoadError` on any problem; never returns a partial bundle.
    """
    resolved = Path(base_dir) if base_dir is not None else settings.roadmap_mapping_dir
    return _load_bundle(resolved, strict=strict)


def get_mappings(force_reload: bool = False) -> MappingBundle:
    """Return the cached bundle for the configured directory, loading it on the
    first call (or when `force_reload=True`). Lazy by design -- see module docstring."""
    global _CACHE
    if _CACHE is None or force_reload:
        _CACHE = load_mappings()
    return _CACHE


def reload_mappings() -> MappingBundle:
    """Force a cache refresh and return the freshly-loaded bundle."""
    return get_mappings(force_reload=True)
