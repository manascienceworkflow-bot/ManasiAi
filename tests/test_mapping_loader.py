import sys
from pathlib import Path

sys.path.append(str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
import pytest  # noqa: E402

from app.roadmap import mapping_loader  # noqa: E402
from app.roadmap.mapping_loader import (  # noqa: E402
    MappingLoadError,
    load_mappings,
)
from app.roadmap.mapping_models import ModalityRow  # noqa: E402

# The layout of the real files: a merged title banner on row 1, two blank rows,
# the header on row 4, data from row 5 on. The default header row here is 4 so
# fixtures mirror production; tests that probe header discovery override it.
DEFAULT_HEADER = ("Domain", "Track", "Suggested modalities (in order)", "Relevance")
DEFAULT_DATA = [
    ("A. Sensory Processing", "Spine", "MNRI", "Primary"),
    (None, None, "Feldenkrais", "Secondary"),
    ("B. Auditory & Language", "Spine", "Tomatis", "Primary"),
]


def _write_xlsx(
    path,
    *,
    sheet="Neurodivergent Path",
    banner="ND Path - Domain to Modality",
    header_row=4,
    header=DEFAULT_HEADER,
    data=DEFAULT_DATA,
):
    """Build a mapping workbook that mirrors the real banner+blanks+header layout.

    `header_row` is the 1-based row the header lands on (rows above it get the
    banner on row 1 and blanks elsewhere). `header=None` writes no header row at
    all (data still lands at header_row+1) -- used for the header-not-found case.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    if banner is not None:
        ws.cell(row=1, column=1, value=banner)
    if header is not None:
        for col, value in enumerate(header, start=1):
            ws.cell(row=header_row, column=col, value=value)
    for r, row in enumerate(data, start=header_row + 1):
        for col, value in enumerate(row, start=1):
            if value is not None:
                ws.cell(row=r, column=col, value=value)
    wb.save(path)
    wb.close()
    return path


def _nd_and_nt(dir_, **nd_kwargs):
    """Write a valid NT file plus an ND file with the given overrides, so a test
    can focus on ND while `load_mappings(base_dir=...)` still finds both."""
    _write_xlsx(dir_ / "Neurodivergent_map.xlsx", sheet="Neurodivergent Path", **nd_kwargs)
    _write_xlsx(
        dir_ / "Neurotypical_map.xlsx",
        sheet="Neurotypical Path",
        banner="Acquired / Situational Path",
        data=[("A. Attention", "Spine", "Arrowsmith", "Primary")],
    )


# --------------------------------------------------------------------------- #
# Happy path / unit behaviour
# --------------------------------------------------------------------------- #


def test_loads_both_files(tmp_path):
    _nd_and_nt(tmp_path)
    bundle = load_mappings(base_dir=tmp_path)
    assert bundle.neurodivergent.source == "ND"
    assert bundle.neurotypical.source == "NT"
    assert bundle.neurodivergent.row_count == 3
    assert bundle.neurotypical.row_count == 1


def test_header_discovered_below_banner(tmp_path):
    _nd_and_nt(tmp_path, header_row=4)
    ds = load_mappings(base_dir=tmp_path).neurodivergent
    assert ds.header_row == 4
    assert ds.rows[0].source_row == 5


def test_header_discovered_at_deeper_offset(tmp_path):
    # Header need not be row 4 -- discovery, not assumption.
    _nd_and_nt(tmp_path, header_row=7)
    ds = load_mappings(base_dir=tmp_path).neurodivergent
    assert ds.header_row == 7
    assert ds.rows[0].source_row == 8


def test_lowercase_modalities_column_matches(tmp_path):
    # The real file spells it lowercase; extra spaces too. Must still match.
    header = ("domain", "TRACK", "Suggested   modalities (in order)", "relevance")
    _nd_and_nt(tmp_path, header=header)
    ds = load_mappings(base_dir=tmp_path).neurodivergent
    assert ds.rows[0].modality == "MNRI"


def test_sparse_domain_track_preserved_as_none(tmp_path):
    _nd_and_nt(tmp_path)
    ds = load_mappings(base_dir=tmp_path).neurodivergent
    assert ds.rows[1].domain is None
    assert ds.rows[1].track is None
    assert ds.rows[1].modality == "Feldenkrais"


def test_values_are_verbatim(tmp_path):
    _nd_and_nt(tmp_path, data=[("  A. Odd  ", "Spine", "MNRI ", "Primary")])
    ds = load_mappings(base_dir=tmp_path).neurodivergent
    # Whitespace and case are untouched -- no strip, no re-case.
    assert ds.rows[0].domain == "  A. Odd  "
    assert ds.rows[0].modality == "MNRI "


def test_blank_interior_row_skipped(tmp_path):
    data = [
        ("A", "Spine", "MNRI", "Primary"),
        (None, None, None, None),
        ("B", "Spine", "Tomatis", "Primary"),
    ]
    _nd_and_nt(tmp_path, data=data)
    ds = load_mappings(base_dir=tmp_path).neurodivergent
    assert [r.modality for r in ds.rows] == ["MNRI", "Tomatis"]


def test_extra_column_ignored_not_error(tmp_path):
    header = DEFAULT_HEADER + ("Notes",)
    data = [("A", "Spine", "MNRI", "Primary", "some note")]
    _nd_and_nt(tmp_path, header=header, data=data)
    ds = load_mappings(base_dir=tmp_path).neurodivergent
    assert ds.rows[0].modality == "MNRI"


def test_single_sheet_autoselected_despite_unexpected_name(tmp_path):
    _nd_and_nt(tmp_path)
    # Rewrite ND with a single, differently-named sheet.
    _write_xlsx(tmp_path / "Neurodivergent_map.xlsx", sheet="Whatever")
    ds = load_mappings(base_dir=tmp_path).neurodivergent
    assert ds.sheet_name == "Whatever"


def test_models_are_frozen(tmp_path):
    _nd_and_nt(tmp_path)
    row = load_mappings(base_dir=tmp_path).neurodivergent.rows[0]
    with pytest.raises(Exception):
        row.modality = "changed"


def test_row_reprs_match_expected(tmp_path):
    _nd_and_nt(tmp_path)
    ds = load_mappings(base_dir=tmp_path).neurodivergent
    assert ds.rows[0] == ModalityRow(
        domain="A. Sensory Processing",
        track="Spine",
        modality="MNRI",
        relevance="Primary",
        source_row=5,
    )


# --------------------------------------------------------------------------- #
# Trailing-blank sentinel / large workbook
# --------------------------------------------------------------------------- #


def test_trailing_blank_sentinel_stops_scan(tmp_path):
    # Two data rows, then a big gap, then a stray row far below the sentinel.
    data = [("A", "Spine", "MNRI", "Primary"), ("B", "Spine", "Tomatis", "Primary")]
    _nd_and_nt(tmp_path, data=data)
    # Manually append a row ~200 rows below -- beyond MAX_TRAILING_BLANK_ROWS (50).
    path = tmp_path / "Neurodivergent_map.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb["Neurodivergent Path"]
    ws.cell(row=250, column=3, value="OrphanModality")
    wb.save(path)
    wb.close()
    ds = load_mappings(base_dir=tmp_path).neurodivergent
    # The orphan is past the sentinel and must not appear.
    assert [r.modality for r in ds.rows] == ["MNRI", "Tomatis"]


# --------------------------------------------------------------------------- #
# Non-strict mode
# --------------------------------------------------------------------------- #


def test_non_strict_drops_row_without_modality_into_warnings(tmp_path):
    data = [
        ("A", "Spine", "MNRI", "Primary"),
        ("B (no modality)", "Spine", None, "Primary"),
    ]
    _nd_and_nt(tmp_path, data=data)
    ds = load_mappings(base_dir=tmp_path).neurodivergent
    assert [r.modality for r in ds.rows] == ["MNRI"]
    assert any("modality" in w for w in ds.warnings)


# --------------------------------------------------------------------------- #
# Failure matrix -- each raises MappingLoadError with a specific code
# --------------------------------------------------------------------------- #


def test_missing_file(tmp_path):
    # Only NT exists; ND is absent.
    _write_xlsx(tmp_path / "Neurotypical_map.xlsx", sheet="Neurotypical Path")
    with pytest.raises(MappingLoadError) as exc:
        load_mappings(base_dir=tmp_path)
    assert exc.value.code == "file_not_found"
    assert exc.value.source == "ND"


def test_path_is_directory(tmp_path):
    (tmp_path / "Neurodivergent_map.xlsx").mkdir()
    _write_xlsx(tmp_path / "Neurotypical_map.xlsx", sheet="Neurotypical Path")
    with pytest.raises(MappingLoadError) as exc:
        load_mappings(base_dir=tmp_path)
    assert exc.value.code == "not_a_file"


def test_corrupt_workbook(tmp_path):
    (tmp_path / "Neurodivergent_map.xlsx").write_bytes(b"not a real xlsx")
    _write_xlsx(tmp_path / "Neurotypical_map.xlsx", sheet="Neurotypical Path")
    with pytest.raises(MappingLoadError) as exc:
        load_mappings(base_dir=tmp_path)
    assert exc.value.code == "workbook_unreadable"


def test_missing_worksheet(tmp_path):
    # ND workbook with TWO sheets, neither the expected name.
    wb = openpyxl.Workbook()
    wb.active.title = "SheetOne"
    wb.create_sheet("SheetTwo")
    wb.save(tmp_path / "Neurodivergent_map.xlsx")
    wb.close()
    _write_xlsx(tmp_path / "Neurotypical_map.xlsx", sheet="Neurotypical Path")
    with pytest.raises(MappingLoadError) as exc:
        load_mappings(base_dir=tmp_path)
    assert exc.value.code == "missing_worksheet"


def test_empty_worksheet(tmp_path):
    _write_xlsx(
        tmp_path / "Neurodivergent_map.xlsx",
        sheet="Neurodivergent Path",
        banner=None,
        header=None,
        data=[],
    )
    _write_xlsx(tmp_path / "Neurotypical_map.xlsx", sheet="Neurotypical Path")
    with pytest.raises(MappingLoadError) as exc:
        load_mappings(base_dir=tmp_path)
    assert exc.value.code == "empty_worksheet"


def test_header_not_found(tmp_path):
    # Only a banner, no header row anywhere.
    _write_xlsx(
        tmp_path / "Neurodivergent_map.xlsx",
        sheet="Neurodivergent Path",
        header=None,
        data=[("just", "some", "stray", "data")],
    )
    _write_xlsx(tmp_path / "Neurotypical_map.xlsx", sheet="Neurotypical Path")
    with pytest.raises(MappingLoadError) as exc:
        load_mappings(base_dir=tmp_path)
    assert exc.value.code == "header_not_found"


def test_missing_column(tmp_path):
    header = ("Domain", "Track", "Suggested modalities (in order)")  # no Relevance
    _nd_and_nt(tmp_path, header=header, data=[("A", "Spine", "MNRI")])
    with pytest.raises(MappingLoadError) as exc:
        load_mappings(base_dir=tmp_path)
    assert exc.value.code == "missing_column"
    assert exc.value.field == "Relevance"


def test_duplicate_column(tmp_path):
    header = ("Domain", "Domain", "Suggested modalities (in order)", "Relevance", "Track")
    _nd_and_nt(tmp_path, header=header, data=[("A", "A2", "MNRI", "Primary", "Spine")])
    with pytest.raises(MappingLoadError) as exc:
        load_mappings(base_dir=tmp_path)
    assert exc.value.code == "duplicate_column"
    assert exc.value.field == "Domain"


def test_no_data_rows(tmp_path):
    _nd_and_nt(tmp_path, data=[])
    with pytest.raises(MappingLoadError) as exc:
        load_mappings(base_dir=tmp_path)
    assert exc.value.code == "no_data_rows"


def test_nd_ok_but_nt_missing_raises_no_partial(tmp_path):
    _write_xlsx(tmp_path / "Neurodivergent_map.xlsx", sheet="Neurodivergent Path")
    with pytest.raises(MappingLoadError) as exc:
        load_mappings(base_dir=tmp_path)
    assert exc.value.code == "file_not_found"
    assert exc.value.source == "NT"


# --------------------------------------------------------------------------- #
# Cache behaviour
# --------------------------------------------------------------------------- #


def test_get_mappings_caches_and_reloads(monkeypatch, tmp_path):
    _nd_and_nt(tmp_path)
    monkeypatch.setattr(mapping_loader.settings, "roadmap_mapping_dir", tmp_path)
    mapping_loader._CACHE = None
    first = mapping_loader.get_mappings()
    assert mapping_loader.get_mappings() is first  # cached identity
    forced = mapping_loader.reload_mappings()
    assert forced is not first  # fresh object after reload
    mapping_loader._CACHE = None  # leave module state clean


# --------------------------------------------------------------------------- #
# Real-file regression -- loads the actual data/roadmap/*.xlsx
# --------------------------------------------------------------------------- #


def test_real_files_load():
    bundle = load_mappings()
    assert bundle.neurodivergent.sheet_name == "Neurodivergent Path"
    assert bundle.neurotypical.sheet_name == "Neurotypical Path"
    assert bundle.neurodivergent.row_count > 0
    assert bundle.neurotypical.row_count > 0
    assert bundle.neurodivergent.rows[0] == ModalityRow(
        domain="A. Sensory Processing",
        track="Spine",
        modality="MNRI",
        relevance="Primary",
        source_row=5,
    )
    # A continuation row keeps its merged-cell None verbatim.
    assert bundle.neurodivergent.rows[1].domain is None


def test_real_files_not_mutated():
    from app.config import settings

    nd = settings.roadmap_mapping_dir / "Neurodivergent_map.xlsx"
    before = nd.stat().st_mtime_ns, nd.stat().st_size
    load_mappings()
    after = nd.stat().st_mtime_ns, nd.stat().st_size
    assert before == after
